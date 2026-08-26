"""Case specification: everything needed to value one policy, serializable to
JSON, plus an extractor that builds a case from an existing InsuriShield
workbook."""
import json, datetime as dt
from dataclasses import dataclass, field, asdict
from .mortality import age_alb

DATEFMT = '%Y-%m-%d'

def _d(x):
    if x is None or isinstance(x, dt.date) and not isinstance(x, dt.datetime): return x
    if isinstance(x, dt.datetime): return x.date()
    return dt.datetime.strptime(x, DATEFMT).date()

@dataclass
class Case:
    name: str
    face: float
    gender: str                    # 'Male' / 'Female'
    smoker: str                    # 'Non-Smoker' / 'Smoker'
    dob: object                    # date or None (falls back to anniversary-anchored age)
    policy_date: object            # first monthiversary anchor
    id_date: object                # illustration date (ledger basis date)
    vd: object                     # valuation date
    maturity_age: int
    illustration_mode: str         # 'Annual' / 'Quarterly' / ...
    av_at_id: float
    ledger: dict                   # {py(int): {prem,ndb,av,csv,gcr,ngcr,popc,ppc,puc,popcat,popcat_t}}
    projection_crediting: float    # AV crediting for the forward roll (analyst assumption)
    ledger_crediting: float = None # crediting inside COI backsolve (default: NGCR)
    funding: dict = field(default_factory=dict)   # {py: 'Illustrated'|'Optimize'|'Custom'}
    custom_premiums: dict = field(default_factory=dict)  # {date: amount}
    coi_overrides: dict = field(default_factory=dict)    # {py: rate} e.g. manually-adjusted years
    mi: float = 0.005              # mortality improvement factor
    survival_table: str = '2015 ALB'
    ndb_lag_months: int = 2
    health_type: str = 'Mean LE50' # or 'Mortality Multiplier'
    health_value: float = 100.0
    valuation_type: str = 'IRR'    # or 'Purchase Price'
    valuation_value: float = 15.0
    n_schedule_months: int = None  # override premium-schedule length (rows)
    insured_name: str = None       # client / insured full name
    payment_frequency: str = 'Monthly'  # buyer payment timing from VD: Monthly/Quarterly/Annual
    illustration_name: str = None  # source illustration file or title
    le_date: object = None         # date of the LE report; when set, the Mean LE50
                                   # is solved AS OF this date and aged to the VD
    le_aging: str = 'condition'    # how an LE report is aged from le_date to VD:
                                   # 'condition' (default, Colva convention) - solve the
                                   #   multiplier at the report date and CONDITION that
                                   #   curve on survival to the VD (S(k+t)/S(k));
                                   # 'rebuild' (InsuriShield classic) - keep the solved
                                   #   multiplier but rebuild the q-series at the VD's
                                   #   attained age (re-anchors select period; slower decay)
    current_year_premium_due: bool = False  # buyer owes the current policy-year
                                   # premium at close (default: seller paid it)
    optimize_basis: str = 'CSV'    # lapse/optimization basis: 'CSV' (default,
                                   # InsuriShield convention) or 'AV' for products
                                   # whose lapse test is on the account value
    nlg_requirement: dict = None   # premium-requirement no-lapse guarantee:
                                   # {'annual': $, 'to_age': age} - cumulative premiums
                                   # must meet annual x years; CSV floor waived while active
    source_runs: list = field(default_factory=list)  # illustration runs considered:
                                   # [{source, label, date, chosen, reason}]
    survivorship: bool = False     # second-to-die policy: price on the joint
                                   # last-survivor curve 1-(1-S1)(1-S2)
    insured2: dict = None          # second insured for survivorship policies:
                                   # {name, gender, smoker, dob 'YYYY-MM-DD',
                                   #  deceased: bool, health_type, health_value,
                                   #  le_date 'YYYY-MM-DD'} - health defaults to
                                   # Mortality Multiplier 100 when absent
    nlg: dict = None               # Lapse Protection rider data for NLG policies:
                                   # {contract_date, issue_age, premium_load, monthly_charge,
                                   #  interest: [[through_yr, rate],...], coi_per_1000: {age: rate},
                                   #  fund_at_vd, ndb}

    # ---------- serialization ----------
    def to_json(self, path):
        d = asdict(self)
        for k in ('dob','policy_date','id_date','vd','le_date'):
            if d.get(k) is not None: d[k] = _d(d[k]).strftime(DATEFMT)
        d['ledger'] = {str(k): v for k, v in self.ledger.items()}
        d['funding'] = {str(k): v for k, v in self.funding.items()}
        d['coi_overrides'] = {str(k): v for k, v in self.coi_overrides.items()}
        d['custom_premiums'] = {(_d(k).strftime(DATEFMT)): v for k, v in self.custom_premiums.items()}
        if d.get('nlg') and d['nlg'].get('contract_date') is not None:
            cd = d['nlg']['contract_date']
            if not isinstance(cd, str):
                d['nlg']['contract_date'] = _d(cd).strftime(DATEFMT)
        with open(path, 'w') as f: json.dump(d, f, indent=1)

    @classmethod
    def from_json(cls, path):
        d = json.load(open(path))
        for k in ('dob','policy_date','id_date','vd','le_date'):
            if d.get(k) is not None: d[k] = dt.datetime.strptime(d[k], DATEFMT).date()
        d['ledger'] = {int(k): v for k, v in d['ledger'].items()}
        d['funding'] = {int(k): v for k, v in d['funding'].items()}
        d['coi_overrides'] = {int(k): v for k, v in d['coi_overrides'].items()}
        d['custom_premiums'] = {dt.datetime.strptime(k, DATEFMT).date(): v
                                for k, v in d['custom_premiums'].items()}
        return cls(**d)

    @property
    def effective_dob(self):
        """DOB used for mortality; when unknown, anchored to the policy
        monthiversary at the ledger issue age (reference behavior)."""
        if self.dob is not None: return _d(self.dob)
        raise ValueError('dob required (set an anniversary-anchored synthetic DOB)')


def extract_case(xlsx_path, name, projection_crediting, dob_override=None):
    """Build a Case from an existing InsuriShield workbook (values only)."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    pi = {r[1].value: r[2].value for r in wb['Policy Inputs'].iter_rows(min_row=2) if r[1].value}
    ps = {r[1].value: r[2].value for r in wb['Pricing Settings'].iter_rows(min_row=2) if r[1].value}
    led = {}
    for r in wb['Ledger'].iter_rows(min_row=2, values_only=True):
        if r[0] is None: continue
        led[int(r[0])] = dict(prem=r[3], ndb=r[4], av=r[5], csv=r[6], gcr=r[7], ngcr=r[8],
                              popc=r[9], ppc=r[10], puc=r[11], popcat=r[12], popcat_t=r[13])
    # funding plan + custom premiums + schedule length from OP sheet
    funding, custom = {}, {}
    n_rows = 0
    op_start = None
    coi_given = {}
    for r in wb['Optimized Premiums'].iter_rows(min_row=2, values_only=True):
        if r[0] is None: break
        n_rows += 1
        py, mode, start = int(r[1]), r[5], r[3]
        if op_start is None: op_start = start
        fm = {'Pay as Illustrated': 'Illustrated'}.get(mode, mode)
        funding[py] = fm
        if fm == 'Custom' and r[11]:
            custom[start.date()] = float(r[11])
        if r[13] is not None: coi_given[py] = float(r[13])
    # manually-adjusted COI years from the COI Rates sheet
    overrides = {}
    first_rate_py = min(coi_given) if coi_given else 1
    i = 0
    for r in wb['COI Rates'].iter_rows(min_row=3, values_only=True):
        if r[0] is None: break
        if r[6] and 'Manual' in str(r[6]) and r[5] is not None:
            overrides[first_rate_py + i] = float(r[5])
        i += 1
    vc = wb['Valuation Calculator']
    dob = pi.get('Date of Birth') or ps.get('Insured DOB (Given)') or dob_override
    case = Case(
        name=name, face=float(pi['Face Amount']), gender=pi['Gender'],
        smoker=pi['Smoking Status'],
        dob=_d(dob) if dob else None,
        policy_date=_d(pi['Policy Date']) if pi.get('Policy Date') else _d(op_start),
        id_date=_d(ps['Illustration Date (ID)']), vd=_d(ps['Valuation Date (VD)']),
        maturity_age=int(pi.get('Maturity Age') or 121),
        illustration_mode=pi.get('Illustration Mode') or 'Annual',
        av_at_id=float(pi.get('AV at ID') or 0.0),
        ledger=led, projection_crediting=projection_crediting,
        funding=funding, custom_premiums=custom, coi_overrides=overrides,
        mi=float(ps.get('Mortality Improvement Factor') or 0.005),
        ndb_lag_months=int(ps.get('NDB Collection Lag (Months)') or 2),
        health_type=vc['C9'].value, health_value=float(vc['C10'].value),
        valuation_type=vc['C11'].value, valuation_value=float(vc['C12'].value),
        n_schedule_months=n_rows,
        insured_name=pi.get('Name'),
        illustration_name=pi.get('Illustration Name'),
    )
    return case
