"""Excel workbook generator: produces a self-contained valuation workbook per
case, architecturally mirroring the InsuriShield model:

- static engine-computed sheets: Policy Inputs, Pricing Settings, Ledger,
  COI Rates, Optimized Premiums, OP (Annual), SC-MM, SC-LE
- live sheets: Valuation Calculator (user inputs w/ dropdowns), Admin,
  Results — so LE50 / mortality multiplier / IRR / purchase price can be
  changed in Excel and the price recomputes with no code.
"""
import numpy as np, datetime as dt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from .mortality import (annual_q_series, survival_curve, solve_mm_for_le,
                        mean_le, median_le, add_months, age_alb, mm_grid)
from .runner import run_case

F_HDR = Font(name='Arial', bold=True, size=11)
F_TXT = Font(name='Arial', size=10)
F_INP = Font(name='Arial', size=10, color='0000FF')
F_TITLE = Font(name='Arial', bold=True, size=14)
FILL_HDR = PatternFill('solid', fgColor='D9E1F2')
FILL_INP = PatternFill('solid', fgColor='FFFF00')
THIN = Border(*[Side(style='thin', color='BFBFBF')]*4)

def _set(ws, cell, value, font=F_TXT, fmt=None, fill=None):
    c = ws[cell]; c.value = value; c.font = font
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    return c

def build_workbook(case, out_path, res=None, le_grid_max=None, sens=None,
                   completeness=None):
    if res is None:
        res = run_case(case)
    rows = res['schedule']; qa = res['qa']
    vd_idx = next(i for i, r in enumerate(rows) if r['start'] == case.vd)
    n_op = len(rows)
    n_surv = len(qa)                       # months after VD in grids
    wb = openpyxl.Workbook()

    # ---------------- grids ----------------
    mms, Smm = mm_grid(qa, 50, 1500, 25)
    grid_dates = [case.vd] + [add_months(case.vd, t+1) for t in range(n_surv)]
    n_grid = len(grid_dates)
    # LE grid - windowed around the case's LE to keep workbook memory/size sane
    le_cap = int(np.floor(mean_le(survival_curve(qa, 25.0))))
    if le_grid_max is None:
        center = int(round(res.get('mean_le') or 120))
        if getattr(case, 'health_type', '') == 'Mean LE50':
            center = int(round(case.health_value))
        le_lo = max(5, center - 90)
        le_grid_max = min(le_cap, center + 150)
    else:
        le_lo = 5
        le_grid_max = min(le_grid_max, le_cap)
    les = list(range(le_lo, le_grid_max+1))
    le_mms = [solve_mm_for_le(qa, le) for le in les]
    Sle = np.column_stack([survival_curve(qa, m) for m in le_mms])

    def write_grid(ws, axis, r2_label_vals, r3_label_vals, S, r1_label):
        ws.cell(1, 1, r1_label)
        ws.cell(2, 1, r2_label_vals[0]); ws.cell(3, 1, r3_label_vals[0])
        for j, a in enumerate(axis):
            ws.cell(1, 2+j, a)
            ws.cell(2, 2+j, round(r2_label_vals[1][j], 1))
            ws.cell(3, 2+j, round(r3_label_vals[1][j], 1))
        for i, d in enumerate(grid_dates):
            ws.cell(4+i, 1, d)
        for j in range(S.shape[1]):
            col = S[:, j]
            for i in range(n_grid):
                ws.cell(4+i, 2+j, round(float(col[i]), 10))

    ws_mm = wb.create_sheet('SC-MM')
    mean_mm = [mean_le(Smm[:, j]) for j in range(len(mms))]
    med_mm = [median_le(Smm[:, j]) for j in range(len(mms))]
    write_grid(ws_mm, mms, ('Mean LE', mean_mm), ('Median LE', med_mm), Smm, 'MM')
    ws_le = wb.create_sheet('SC-LE')
    med_le_vals = [median_le(Sle[:, j]) for j in range(len(les))]
    write_grid(ws_le, les, ('Median LE', med_le_vals), ('MM', le_mms), Sle, 'Mean LE')
    ws_mm.sheet_state = 'hidden'; ws_le.sheet_state = 'hidden'

    # ---------------- static input sheets ----------------
    ws = wb.create_sheet('Policy Inputs')
    pi_rows = [('Reference','Case Name',case.name),
        ('Reference','Insured Name',getattr(case,'insured_name',None)),
        ('Reference','Source Illustration',getattr(case,'illustration_name',None)),
        ('Policy Setup','Policy Date',case.policy_date),
        ('Policy Setup','Policy Type','Universal Life'),
        ('Insured','Gender',case.gender), ('Insured','Smoking Status',case.smoker),
        ('Insured','Date of Birth',case.dob),
        ('Pricing Basics','Face Amount',case.face),
        ('Pricing Basics','Maturity Age',case.maturity_age),
        ('Pricing Basics','AV at ID',case.av_at_id),
        ('Pricing Basics','Illustration Mode',case.illustration_mode)]
    for k, run in enumerate(getattr(case, 'source_runs', None) or [], 1):
        mark = 'CHOSEN' if run.get('chosen') else 'considered'
        desc = f"{run.get('label','')} — {run.get('source','')}"
        if run.get('date'): desc += f" ({run['date']})"
        if run.get('chosen') and run.get('reason'): desc += f" | {run['reason']}"
        pi_rows.append(('Illustrations', f'Run {k}: {mark}', desc))
    _set(ws,'A1','Section',F_HDR); _set(ws,'B1','Input Field',F_HDR); _set(ws,'C1','Field Value',F_HDR)
    for i,(a,b,c) in enumerate(pi_rows,2):
        _set(ws,f'A{i}',a); _set(ws,f'B{i}',b); v=_set(ws,f'C{i}',c,F_INP)
        if isinstance(c, dt.date): v.number_format='yyyy-mm-dd'
    ws.column_dimensions['A'].width=16; ws.column_dimensions['B'].width=22; ws.column_dimensions['C'].width=26

    ws = wb.create_sheet('Pricing Settings')
    ps_rows=[('Key Dates','Illustration Date (ID)',case.id_date),
        ('Key Dates','Valuation Date (VD)',case.vd),
        ('Key Dates','Insured DOB (Used)',case.effective_dob),
        ('Valuation','Survival Table',case.survival_table),
        ('Valuation','Mortality Improvement Factor',case.mi),
        ('Valuation','LE Aging Convention',
            'Survival-conditioned (Colva)' if getattr(case,'le_aging','condition') == 'condition'
            else 'Rebuild at VD (InsuriShield)'),
        ('Valuation','NDB Collection Lag (Months)',case.ndb_lag_months),
        ('Crediting','Projection Crediting Rate',case.projection_crediting),
        ('Crediting','Ledger (Backsolve) Crediting Rate',
            case.ledger_crediting if case.ledger_crediting is not None
            else (case.ledger[min(case.ledger)].get('ngcr') or 0.0)),
        ('Optimization','Premium Optimization Mode','Monthly'),
        ('Optimization','AV/CSV Buffer','1 Month of COI'),
        ('Optimization','Optimize Basis',
            'Account Value' if getattr(case,'optimize_basis','CSV') == 'AV'
            else 'Cash Surrender Value')]
    if getattr(case, 'survivorship', False):
        i2 = getattr(case, 'insured2', None) or {}
        ps_rows.append(('Valuation','Survivorship',
            ('Joint last-survivor curve; insured 2: ' + str(i2.get('name','?')))
            if i2.get('health_value') is not None and not i2.get('deceased')
            else ('Single life (insured 2 deceased: ' + str(i2.get('name','?')) + ')')
            if i2.get('deceased')
            else 'Single-LE convention (LE on the longer-lived insured)'))
    if getattr(case, 'nlg_requirement', None):
        ps_rows.append(('Optimization','NLG Premium Requirement',
            f"${case.nlg_requirement.get('annual',0):,.2f}/yr to age "
            f"{case.nlg_requirement.get('to_age','?')} (CSV floor waived while active)"))
    _set(ws,'A1','Section',F_HDR); _set(ws,'B1','Input Field',F_HDR); _set(ws,'C1','Field Value',F_HDR)
    for i,(a,b,c) in enumerate(ps_rows,2):
        _set(ws,f'A{i}',a); _set(ws,f'B{i}',b); v=_set(ws,f'C{i}',c,F_INP)
        if isinstance(c, dt.date): v.number_format='yyyy-mm-dd'
        if 'Rate' in b or 'Factor' in b: v.number_format='0.000%'
    ws.column_dimensions['A'].width=14; ws.column_dimensions['B'].width=30; ws.column_dimensions['C'].width=18

    ws = wb.create_sheet('Ledger')
    hdr=['Policy Year','Premium Outlay','NDB','Account Value','Cash Surrender Value',
         'GCR','NGCR','POPC','PPC','PUC','POPCAT','POPCAT Target']
    for j,h in enumerate(hdr,1): _set(ws,f'{get_column_letter(j)}1',h,F_HDR,fill=FILL_HDR)
    for i,py in enumerate(sorted(case.ledger),2):
        lp=case.ledger[py]
        vals=[py,lp.get('prem'),lp.get('ndb'),lp.get('av'),lp.get('csv'),lp.get('gcr'),
              lp.get('ngcr'),lp.get('popc'),lp.get('ppc'),lp.get('puc'),lp.get('popcat'),lp.get('popcat_t')]
        for j,v in enumerate(vals,1):
            c=ws.cell(i,j,v); c.font=F_TXT
            if j in (2,3,4,5,12): c.number_format='#,##0'
            if j in (6,7,8): c.number_format='0.00%'
    for j in range(1,13): ws.column_dimensions[get_column_letter(j)].width=13

    ws = wb.create_sheet('COI Rates')
    _set(ws,'A1','Policy Year',F_HDR,fill=FILL_HDR); _set(ws,'B1','COI Rate (annual)',F_HDR,fill=FILL_HDR)
    _set(ws,'C1','Source',F_HDR,fill=FILL_HDR)
    for i,py in enumerate(sorted(res['coi_rates']),2):
        ws.cell(i,1,py).font=F_TXT
        c=ws.cell(i,2,res['coi_rates'][py]); c.font=F_TXT; c.number_format='0.00000000'
        ws.cell(i,3,'Manually adjusted' if py in case.coi_overrides else
                'Backsolved from ledger AV').font=F_TXT
    ws.column_dimensions['A'].width=12; ws.column_dimensions['B'].width=16; ws.column_dimensions['C'].width=26

    # ---------------- Optimized Premiums ----------------
    ws = wb.create_sheet('Optimized Premiums')
    hdr=['Period','Policy Year','Month','Start Date','End Date','Funding Approach',
         'Paid By','Beginning AV','Beginning CSV','Optimized Premium','Net Death Benefit',
         'COI Rate','COI Amount','Ending AV','Ending CSV']
    for j,h in enumerate(hdr,1): _set(ws,f'{get_column_letter(j)}1',h,F_HDR,fill=FILL_HDR)
    for i,r in enumerate(rows,2):
        vals=[i-2-vd_idx, r['py'], r['m'], r['start'], r['end'], r['mode'],
              'Seller' if i-2-vd_idx<0 else 'Buyer', r['beg'], r['beg_csv'],
              r['prem'], r['ndb'], r['rate'], r['coi'], r['end_av'], r['end_csv']]
        for j,v in enumerate(vals,1):
            c=ws.cell(i,j,v); c.font=F_TXT
            if j in (4,5): c.number_format='yyyy-mm-dd'
            if j in (8,9,10,11,13,14,15): c.number_format='#,##0.00'
            if j==12: c.number_format='0.00000000'
    for j in range(1,16): ws.column_dimensions[get_column_letter(j)].width=13

    # ---------------- OP (Annual) ----------------
    ws = wb.create_sheet('OP (Annual)')
    hdr=['Policy Year','Start Date','Annual Premium','Paid By','NDB']
    for j,h in enumerate(hdr,1): _set(ws,f'{get_column_letter(j)}1',h,F_HDR,fill=FILL_HDR)
    i=2
    for py in sorted(set(r['py'] for r in rows)):
        yr=[r for r in rows if r['py']==py]
        ws.cell(i,1,py).font=F_TXT
        c=ws.cell(i,2,yr[0]['start']); c.font=F_TXT; c.number_format='yyyy-mm-dd'
        c=ws.cell(i,3,sum(r['prem'] for r in yr)); c.font=F_TXT; c.number_format='#,##0.00'
        ws.cell(i,4,'Seller' if yr[-1]['start']<case.vd else 'Buyer').font=F_TXT
        c=ws.cell(i,5,yr[0]['ndb']); c.font=F_TXT; c.number_format='#,##0'
        i+=1
    for j in range(1,6): ws.column_dimensions[get_column_letter(j)].width=15

    # ---------------- Admin ----------------
    wsa = wb.create_sheet('Admin')
    _set(wsa,'A1','Mortality Multiplier'); _set(wsa,'A2','Mean LE50')
    _set(wsa,'A3','IRR'); _set(wsa,'A4','Purchase Price')
    _set(wsa,'A8','MM Input'); wsa['B8']="='Valuation Calculator'!$C$10"
    _set(wsa,'A9','MM Low');  wsa['B9']='=ROUNDDOWN($B$8/25,0)*25'
    _set(wsa,'A10','MM High'); wsa['B10']='=ROUNDUP($B$8/25,0)*25'
    _set(wsa,'A12','VD Row (Results)'); wsa['B12']=vd_idx+2
    _set(wsa,'A13','Last OP Row (Results)'); wsa['B13']=n_op+1
    _set(wsa,'A15','Last Grid Row (Results)'); wsa['B15']=n_grid+1
    _set(wsa,'A16','Breakeven Row'); wsa['B16']='=IFERROR(MATCH(1,Results!$M:$M,0),0)'
    _set(wsa,'A18','XIRR (0.1)')
    wsa['B18']=(f'=IFERROR(XIRR(INDEX(Results!$I:$I,$B$12):INDEX(Results!$I:$I,$B$13+3),'
                f'INDEX(Results!$B:$B,$B$12):INDEX(Results!$B:$B,$B$13+3),0.1),"")')
    _set(wsa,'A19','XIRR (-0.05)')
    wsa['B19']=(f'=IFERROR(XIRR(INDEX(Results!$I:$I,$B$12):INDEX(Results!$I:$I,$B$13+3),'
                f'INDEX(Results!$B:$B,$B$12):INDEX(Results!$B:$B,$B$13+3),-0.05),"")')
    wsa.sheet_state='hidden'

    # ---------------- Results ----------------
    wsr = wb.create_sheet('Results')
    rhdr=['Period','Start Date','Paid By','Purchase Price','S Low','S High','Survival',
          'Probabilistic Cash Flow','PCF Discounted','Cumulative PCF',
          'Total Negative Cashflow','NDB - TNC','BE Flag']
    for j,h in enumerate(rhdr,1): _set(wsr,f'{get_column_letter(j)}1',h,F_HDR,fill=FILL_HDR)
    VC="'Valuation Calculator'"
    n_res = n_grid + 1   # data rows
    for i in range(2, n_res+1):
        t = i-2  # months from schedule start? careful: results rows aligned to OP rows
        # A: period, B: date (values)
        wsr.cell(i,1,i-(vd_idx+2)).font=F_TXT
        if i-2 < n_op:
            d=rows[i-2]['start']
        else:
            d=grid_dates[(i-2)-vd_idx] if (i-2)-vd_idx < len(grid_dates) else None
        if d is None: break
        c=wsr.cell(i,2,d); c.font=F_TXT; c.number_format='yyyy-mm-dd'
        wsr.cell(i,3,'Seller' if i-(vd_idx+2)<0 else ('Buyer' if i-2<n_op else '')).font=F_TXT
        # D: purchase price at VD row in Purchase Price mode
        wsr.cell(i,4,(f'=IF(ROW()>Admin!$B$13+3,"",IF(OR(ROW()<>Admin!$B$12,{VC}!$C$11="IRR"),0,{VC}!$C$12))')).font=F_TXT
        # E/F: survival at low/high MM (or LE column); G: interpolated
        e=(f'=IF($A{i}<0,1,IF({VC}!$C$9="Mortality Multiplier",'
           f"INDEX('SC-MM'!$1:$1048576,MATCH($B{i},'SC-MM'!$A:$A,0),MATCH(Admin!$B$9,'SC-MM'!$1:$1,0)),"
           f"INDEX('SC-LE'!$1:$1048576,MATCH($B{i},'SC-LE'!$A:$A,0),MATCH({VC}!$C$10,'SC-LE'!$1:$1,0))))")
        f_=(f'=IF($A{i}<0,1,IF({VC}!$C$9="Mortality Multiplier",'
           f"INDEX('SC-MM'!$1:$1048576,MATCH($B{i},'SC-MM'!$A:$A,0),MATCH(Admin!$B$10,'SC-MM'!$1:$1,0)),"
           f"INDEX('SC-LE'!$1:$1048576,MATCH($B{i},'SC-LE'!$A:$A,0),MATCH({VC}!$C$10,'SC-LE'!$1:$1,0))))")
        g=(f'=IF($A{i}<0,1,IF({VC}!$C$9="Mortality Multiplier",'
           f'IF(Admin!$B$9=Admin!$B$8,$E{i},((25-(Admin!$B$8-Admin!$B$9))*$E{i}+(25-(Admin!$B$10-Admin!$B$8))*$F{i})/25),'
           f'$E{i}))')
        wsr.cell(i,5,e).font=F_TXT; wsr.cell(i,6,f_).font=F_TXT; wsr.cell(i,7,g).font=F_TXT
        for col in (5,6,7): wsr.cell(i,col).number_format='0.0000000000'
        # H..M only through n_op+3
        if i <= n_op+1+3:
            h=(f'=IF(ROW()>Admin!$B$13+3,"",IF($A{i}<0,0,IF($A{i}<3,0,'
               f"(INDEX($G:$G,ROW()-3)-INDEX($G:$G,ROW()-2))*INDEX('Optimized Premiums'!$K:$K,ROW()-3))"
               f"-$G{i}*(IF(ROW()<=Admin!$B$13,INDEX('Optimized Premiums'!$J:$J,ROW()),0)+$D{i})))")
            wsr.cell(i,8,h).font=F_TXT
            wsr.cell(i,9,(f'=IF(ROW()>Admin!$B$13+3,"",$H{i}/(1+IF({VC}!$C$11="IRR",{VC}!$C$12/100,0))^($A{i}/12))')).font=F_TXT
            wsr.cell(i,10,(f'=IF(ROW()>Admin!$B$13+3,"",IF($A{i}<0,0,SUM($H$2:$H{i})))')).font=F_TXT
            k=(f'=IF(ROW()>Admin!$B$13,"",IF($A{i}<0,0,IF($A{i}=0,MAX(IF({VC}!$C$11="IRR",{VC}!$F$16+0,{VC}!$C$12),0),$K{i-1})'
               f"+IF(ROW()<=Admin!$B$13,INDEX('Optimized Premiums'!$J:$J,ROW()),0)))")
            wsr.cell(i,11,k).font=F_TXT
            wsr.cell(i,12,(f"=IF(ROW()>Admin!$B$13,\"\",INDEX('Optimized Premiums'!$K:$K,ROW())-$K{i})")).font=F_TXT
            wsr.cell(i,13,(f'=IF(ROW()>Admin!$B$13,"",IF(N($L{i})<0,IF(COUNTIF($M$1:$M{i-1},1)=0,1,""),""))')).font=F_TXT
            for col in (8,9,10,11,12): wsr.cell(i,col).number_format='#,##0.00'
    wsr.sheet_state='hidden'

    # ---------------- Valuation Calculator ----------------
    wsv = wb.create_sheet('Valuation Calculator', 0)
    del wb['Sheet']
    _set(wsv,'B2',f'{case.name} — Valuation Model',F_TITLE)
    _set(wsv,'B3','Reproduction engine: survival curves (2015 VBT ALB), COI backsolve, premium optimization, probabilistic valuation',Font(name='Arial',italic=True,size=9))
    _set(wsv,'B6','User Inputs',F_HDR,fill=FILL_HDR); _set(wsv,'E6','Key Pricing Metrics',F_HDR,fill=FILL_HDR); _set(wsv,'H6','Key Case Details',F_HDR,fill=FILL_HDR)
    _set(wsv,'B9','Health Input Variable:'); _set(wsv,'C9',case.health_type,F_INP,fill=FILL_INP)
    wsv['B10']='=$C$9&":"'; wsv['B10'].font=F_TXT
    _set(wsv,'C10',case.health_value,F_INP,fill=FILL_INP)
    _set(wsv,'B11','Valuation Input Variable:'); _set(wsv,'C11',case.valuation_type,F_INP,fill=FILL_INP)
    wsv['B12']='=$C$11&":"'; wsv['B12'].font=F_TXT
    _set(wsv,'C12',case.valuation_value,F_INP,fill=FILL_INP)
    dv1=DataValidation(type='list',formula1='"Mean LE50,Mortality Multiplier"',allow_blank=False)
    dv2=DataValidation(type='list',formula1='"IRR,Purchase Price"',allow_blank=False)
    wsv.add_data_validation(dv1); wsv.add_data_validation(dv2)
    dv1.add('C9'); dv2.add('C11')
    # metrics
    wsv['E9']='=IF($C$9="Mortality Multiplier","Selected Mortality Multiplier:","Selected Mean LE50:")'
    wsv['F9']='=IF($C$9="Mortality Multiplier",CONCATENATE($C$10,"%"),CONCATENATE($C$10," Months"))'
    wsv['E10']='=IF($C$9="Mortality Multiplier","Calculated Mean LE50:","Calculated Mortality Multiplier:")'
    wsv['F10']=('=IF($C$9="Mortality Multiplier",CONCATENATE(FIXED(SUM(Results!$G:$G)-(Admin!$B$12-2)-0.5,1)," Months"),'
                "CONCATENATE(FIXED(INDEX('SC-LE'!$3:$3,MATCH($C$10,'SC-LE'!$1:$1,0)),1),\"%\"))")
    wsv['E11']='=IF($C$11="Purchase Price","Selected Purchase Price:","Selected IRR:")'
    wsv['F11']='=IF($C$11="Purchase Price",CONCATENATE("$",FIXED($C$12,2)),CONCATENATE(FIXED($C$12,2),"%"))'
    wsv['E12']='=IF($C$11="Purchase Price","Calculated IRR:","Calculated Purchase Price:")'
    wsv['F12']=('=IF($C$11="IRR",CONCATENATE("$",FIXED(SUM(Results!$I:$I),2)),'
                'CONCATENATE(FIXED(IF(ROUND(N(Admin!$B$18),4)<>0,N(Admin!$B$18),N(Admin!$B$19))*100,2),"%"))')
    _set(wsv,'E13','Breakeven Risk:')
    wsv['F13']=('=IF(INDEX(Results!$L:$L,Admin!$B$12)<0,1,IF(Admin!$B$16=0,$F$14,'
                'INDEX(Results!$G:$G,Admin!$B$16)))')
    _set(wsv,'E14','Probability of Survival to Maturity:')
    wsv['F14']='=INDEX(Results!$G:$G,Admin!$B$13+1)'
    _set(wsv,'E16','Purchase Price (numeric):')
    wsv['F16']='=IF($C$11="IRR",SUM(Results!$I:$I),$C$12)'
    wsv['F16'].number_format='#,##0.00'
    for cell in ('E9','E10','E11','E12'): wsv[cell].font=F_TXT
    for cell in ('F9','F10','F11','F12','F13','F14'): wsv[cell].font=Font(name='Arial',bold=True,size=10)
    wsv['F13'].number_format='General'; wsv['F14'].number_format='General'
    # case details
    det=[('Case Name:',case.name),('Gender / Smoking:',f'{case.gender} / {case.smoker}'),
         ('DOB (used):',str(case.effective_dob)),('Face Amount:',case.face),
         ('Valuation Date:',str(case.vd)),('Policy Date:',str(case.policy_date)),
         ('Projection Crediting:',case.projection_crediting)]
    for k,(a,b) in enumerate(det):
        _set(wsv,f'H{9+k}',a); c=_set(wsv,f'I{9+k}',b,Font(name='Arial',bold=True,size=10))
        if a=='Face Amount:': c.number_format='$#,##0'
        if a=='Projection Crediting:': c.number_format='0.00%'
    _set(wsv,'B15','Static Settings',F_HDR,fill=FILL_HDR)
    stat=[('Survival Table:',case.survival_table),
          ('LE grid coverage:',f'{les[0]}-{les[-1]} months (regenerate workbook for LEs outside this range)'),
          ('Mortality Improvement:',case.mi),
          ('NDB Collection Lag:',f'{case.ndb_lag_months} Months'),
          ('Premium Optimization:','Monthly, CSV >= 1 month of COI'),
          ('Note:','Grids & premiums are engine-computed; change C9-C12 freely.')]
    for k,(a,b) in enumerate(stat):
        _set(wsv,f'B{17+k}',a); c=_set(wsv,f'C{17+k}',b)
        if a=='Mortality Improvement:': c.number_format='0.0%'
    # sensitivity / confidence band (static values computed at generation time;
    # the SC-LE / SC-MM grids stay live for interactive use)
    if sens:
        r0 = 25
        _set(wsv,f'B{r0}','Sensitivity — Confidence Band',F_HDR,fill=FILL_HDR)
        _set(wsv,f'B{r0+1}',f"Prices at generation; LE row @ {sens['base_irr']:g}% IRR, "
                            'IRR row @ the base health input. Bold = base case.')
        rr = r0+2
        if sens.get('le'):
            _set(wsv,f'B{rr}','LE (months)',F_HDR)
            for j,(le,p) in enumerate(sens['le']):
                bold = (sens.get('base_le') is not None and le == sens['base_le'])
                _set(wsv,f'{get_column_letter(3+j)}{rr}',le,
                     Font(name='Arial',bold=bold,size=10))
            _set(wsv,f'B{rr+1}',f"Price @ {sens['base_irr']:g}%",F_HDR)
            for j,(le,p) in enumerate(sens['le']):
                bold = (sens.get('base_le') is not None and le == sens['base_le'])
                c=_set(wsv,f'{get_column_letter(3+j)}{rr+1}',p,
                       Font(name='Arial',bold=bold,size=10))
                c.number_format='$#,##0'
            rr += 3
        if sens.get('irr'):
            _set(wsv,f'B{rr}','Target IRR',F_HDR)
            for j,(irr,p) in enumerate(sens['irr']):
                bold = (irr == sens.get('base_irr'))
                c=_set(wsv,f'{get_column_letter(3+j)}{rr}',irr/100.0,
                       Font(name='Arial',bold=bold,size=10))
                c.number_format='0.0%'
            _set(wsv,f'B{rr+1}','Price',F_HDR)
            for j,(irr,p) in enumerate(sens['irr']):
                bold = (irr == sens.get('base_irr'))
                c=_set(wsv,f'{get_column_letter(3+j)}{rr+1}',p,
                       Font(name='Arial',bold=bold,size=10))
                c.number_format='$#,##0'
    # input completeness (green / amber / red source confidence)
    if completeness:
        rc = 25 + (10 if sens else 0)
        FILLS = {'ok': PatternFill('solid', fgColor='D1FAE5'),
                 'warn': PatternFill('solid', fgColor='FEF3C7'),
                 'miss': PatternFill('solid', fgColor='FEE2E2')}
        WORD = {'ok': 'OK', 'warn': 'CHECK', 'miss': 'MISSING'}
        _set(wsv, f'B{rc}', 'Input Completeness', F_HDR, fill=FILL_HDR)
        n_ok = sum(1 for c in completeness if c.get('status') == 'ok')
        _set(wsv, f'B{rc+1}', f'{n_ok} of {len(completeness)} pricing inputs fully '
             'sourced. Green = documented; amber = could be tightened; red = missing.')
        for k, c in enumerate(completeness):
            r = rc + 2 + k
            st = c.get('status', 'warn')
            cell = _set(wsv, f'B{r}', c.get('label', ''), F_HDR)
            v = _set(wsv, f'C{r}', WORD.get(st, st.upper()),
                     Font(name='Arial', bold=True, size=9))
            v.fill = FILLS.get(st, FILLS['warn'])
            _set(wsv, f'E{r}', c.get('source', ''))
            _set(wsv, f'F{r}', c.get('note', ''))
    wb.calculation.fullCalcOnLoad = True   # Excel/LibreOffice recalc on open
    wsv.column_dimensions['A'].width=2
    for col,w in dict(B=26,C=22,D=2,E=30,F=22,G=2,H=20,I=28).items():
        wsv.column_dimensions[col].width=w
    return wb, res
